import csv
import os
import platform
import random
import subprocess
import sys
import simpy

from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple
from alg import BaseRepairAlgorithm
from models import DataBlock, EdgeNode, NodeStatus

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Предупреждение: openpyxl не установлен. Экспорт в Excel будет недоступен.")
    print("Установите: pip install openpyxl")

try:
    import simpy
    import numpy as np
    import matplotlib.pyplot as plt
    from matplotlib.animation import FuncAnimation
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    print("Установите необходимые библиотеки:")
    print("pip install simpy numpy matplotlib")
    sys.exit(1)

class EdgeStorageSystem:
    """Распределённая система хранения данных"""
    
    def __init__(self, env: simpy.Environment, num_nodes: int, 
                 replication_factor: int, config: Dict, metrics_collector):
        self.env = env
        self.num_nodes = num_nodes
        self.replication_factor = replication_factor
        self.config = config
        self.metrics = metrics_collector
        
        self.nodes: List[EdgeNode] = []
        for i in range(num_nodes):
            node = EdgeNode(i, env, config, metrics_collector)
            self.nodes.append(node)
        
        self.block_placement: Dict[int, List[int]] = defaultdict(list)
        self.next_block_id = 0
        
    def _select_nodes_for_replica(self, exclude_nodes: Set[int] = None) -> List[int]:
        if exclude_nodes is None:
            exclude_nodes = set()
        
        available_nodes = [n.node_id for n in self.nodes 
                          if n.status == NodeStatus.ONLINE 
                          and n.node_id not in exclude_nodes]
        
        if len(available_nodes) < self.replication_factor:
            return available_nodes
        
        return random.sample(available_nodes, self.replication_factor)
    
    def write_block(self, data_size: int = 1024) -> Optional[int]:
        block_id = self.next_block_id
        self.next_block_id += 1
        
        block = DataBlock(block_id, data_size, datetime.now())
        target_nodes = self._select_nodes_for_replica()
        
        if not target_nodes:
            self.metrics.record_write_failure(block_id, "no_nodes_available")
            return None
        
        successful_writes = 0
        for node_id in target_nodes:
            if self.nodes[node_id].store_block(block):
                successful_writes += 1
        
        if successful_writes >= self.config.get('min_writes_for_success', 1):
            self.block_placement[block_id] = target_nodes
            self.metrics.record_write_success(block_id, successful_writes, 
                                              len(target_nodes))
            return block_id
        else:
            self.metrics.record_write_failure(block_id, "insufficient_replicas")
            return None
    
    def read_block(self, block_id: int) -> Optional[DataBlock]:
        if block_id not in self.block_placement:
            return None
        
        for node_id in self.block_placement[block_id]:
            if self.nodes[node_id].status == NodeStatus.ONLINE:
                block = self.nodes[node_id].read_block(block_id)
                if block:
                    self.metrics.record_read_success(block_id, node_id)
                    return block
        
        self.metrics.record_read_failure(block_id)
        return None
    
    def repair_degraded_replicas(self, block_id: int):
        """Базовая операция восстановления реплик"""
        if block_id not in self.block_placement:
            return
        
        current_nodes = set(self.block_placement[block_id])
        online_nodes = {n.node_id for n in self.nodes if n.status == NodeStatus.ONLINE}
        offline_nodes = current_nodes - online_nodes
        
        if not offline_nodes:
            return
        
        needed_replicas = len(offline_nodes)
        available_nodes = [n.node_id for n in self.nodes 
                          if n.status == NodeStatus.ONLINE 
                          and n.node_id not in current_nodes]
        
        for target_node in available_nodes[:needed_replicas]:
            source_nodes = list(current_nodes - offline_nodes)
            if source_nodes:
                source = random.choice(source_nodes)
                block = self.nodes[source].read_block(block_id)
                if block and self.nodes[target_node].store_block(block):
                    self.block_placement[block_id].append(target_node)
                    self.metrics.record_repair_success(block_id, source, target_node)
        
        self.block_placement[block_id] = [n for n in self.block_placement[block_id] 
                                          if n in online_nodes]
        return
    
    def get_availability_score(self) -> float:
        if not self.block_placement:
            return 1.0
        
        total_blocks = len(self.block_placement)
        available_blocks = 0
        
        for block_id, nodes in self.block_placement.items():
            for node_id in nodes:
                if self.nodes[node_id].status == NodeStatus.ONLINE:
                    available_blocks += 1
                    break
        
        return available_blocks / total_blocks if total_blocks > 0 else 1.0
    
    def get_stats(self) -> Dict:
        return {
            'num_nodes': self.num_nodes,
            'online_nodes': sum(1 for n in self.nodes if n.status == NodeStatus.ONLINE),
            'total_blocks': len(self.block_placement),
            'availability': self.get_availability_score(),
            'replication_factor': self.replication_factor,
            'node_stats': [n.get_stats() for n in self.nodes]
        }

class AggressiveEnvironment:
    """Симулятор агрессивных условий внешней среды"""
    
    def __init__(self, env: simpy.Environment, storage_system: EdgeStorageSystem,
                 failure_rate: float, config: Dict, repair_algorithm: BaseRepairAlgorithm,
                 log_callback=None, stop_event=None):
        self.env = env
        self.storage = storage_system
        self.failure_rate = failure_rate
        self.config = config
        self.repair_algorithm = repair_algorithm
        self.log_callback = log_callback
        self.active = True
        self.stop_event = stop_event
        
    def run(self):
        while self.active and (self.stop_event is None or not self.stop_event.is_set()):
            if self.failure_rate > 0:
                time_to_failure = random.expovariate(self.failure_rate)
                yield self.env.timeout(time_to_failure)
                
                if self.active and (self.stop_event is None or not self.stop_event.is_set()):
                    if self.storage.num_nodes > 0:
                        node_idx = random.randint(0, self.storage.num_nodes - 1)
                        node = self.storage.nodes[node_idx]
                        
                        if node.status == NodeStatus.ONLINE:
                            node.fail()
                            self._log(f"[СБОЙ] Узел {node.node_id} отказал в момент {self.env.now:.1f}")
                            self.env.process(self._schedule_recovery(node))
    
    def _schedule_recovery(self, node: EdgeNode):
        detection_delay = random.uniform(
            self.config.get('failure_detection_delay', 1.0),
            self.config.get('failure_detection_delay', 1.0) * 2
        )
        yield self.env.timeout(detection_delay)
        
        if self.active and node.status == NodeStatus.OFFLINE:
            self._log(f"[ВОССТАНОВЛЕНИЕ] Узел {node.node_id} начинает восстановление")
            yield self.env.process(node.recover())
            self._log(f"[ВОССТАНОВЛЕНИЕ] Узел {node.node_id} восстановлен в момент {self.env.now:.1f}")
            
            # Запуск выбранного алгоритма восстановления
            self.env.process(self._run_repair_algorithm(node.node_id))
    
    def _run_repair_algorithm(self, failed_node_id: int):
        """Запуск выбранного алгоритма восстановления"""
        self._log(f"[АЛГОРИТМ] Запуск алгоритма: {self.repair_algorithm.name}")
        yield self.env.process(self.repair_algorithm.repair(
            self.env, self.storage, failed_node_id, self._log
        ))
        self._log(f"[АЛГОРИТМ] Алгоритм {self.repair_algorithm.name} завершил работу")
    
    def _log(self, message: str):
        if self.log_callback:
            self.log_callback(f"[{self.env.now:.1f}] {message}")

class MetricsCollector:
    """Сбор метрик с поддержкой CSV и Excel экспорта"""
    
    def __init__(self):
        self.reset()
    
    def reset(self):
        self.node_failures = []
        self.node_recoveries = []
        self.write_successes = []
        self.write_failures = []
        self.read_successes = []
        self.read_failures = []
        self.repair_successes = []
        self.availability_history = []
        self.health_history = []
        
        self.simulation_start_time = None
        self.simulation_end_time = None
        self.simulation_params = {}
        
    def record_node_failure(self, node_id: int, time: float):
        self.node_failures.append((time, node_id))
        
    def record_node_recovery(self, node_id: int, time: float):
        self.node_recoveries.append((time, node_id))
        
    def record_write_success(self, block_id: int, replicas_written: int, target_replicas: int):
        self.write_successes.append((block_id, replicas_written, target_replicas))
        
    def record_write_failure(self, block_id: int, reason: str):
        self.write_failures.append((block_id, reason))
        
    def record_read_success(self, block_id: int, node_id: int):
        self.read_successes.append((block_id, node_id))
        
    def record_read_failure(self, block_id: int):
        self.read_failures.append((block_id))
        
    def record_repair_success(self, block_id: int, source_node: int, target_node: int):
        self.repair_successes.append((block_id, source_node, target_node))
        
    def record_availability(self, time: float, score: float):
        self.availability_history.append((time, score))
        
    def record_health(self, time: float, online_nodes: int, total_nodes: int):
        self.health_history.append((time, online_nodes, total_nodes))
    
    def set_simulation_params(self, params: Dict):
        self.simulation_params = params
        self.simulation_start_time = datetime.now()
    
    def set_simulation_end_time(self):
        self.simulation_end_time = datetime.now()
        
    def get_summary(self) -> Dict:
        return {
            'total_failures': len(self.node_failures),
            'total_recoveries': len(self.node_recoveries),
            'total_writes_success': len(self.write_successes),
            'total_writes_failed': len(self.write_failures),
            'total_reads_success': len(self.read_successes),
            'total_reads_failed': len(self.read_failures),
            'total_repairs': len(self.repair_successes),
            'avg_replicas_per_write': np.mean([w[1] for w in self.write_successes]) if self.write_successes else 0,
            'write_success_rate': len(self.write_successes) / (len(self.write_successes) + len(self.write_failures)) if (self.write_successes or self.write_failures) else 1.0,
            'avg_availability': np.mean([a[1] for a in self.availability_history]) if self.availability_history else 1.0,
            'min_availability': min([a[1] for a in self.availability_history]) if self.availability_history else 1.0,
            'max_availability': max([a[1] for a in self.availability_history]) if self.availability_history else 1.0,
            'total_blocks_written': len(self.write_successes)
        }
    
    def _open_file_location(self, filepath: str):
        """Открыть папку с файлом в проводнике"""
        try:
            path = os.path.dirname(filepath)
            if not path:  # Если файл сохраняется в текущей директории
                path = os.getcwd()
            
            if os.path.exists(path):
                if platform.system() == 'Windows':
                    os.startfile(path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.run(['open', path], check=False)
                else:  # Linux
                    subprocess.run(['xdg-open', path], check=False)
        except Exception as e:
            # Не критичная ошибка - просто логируем, но не прерываем экспорт
            print(f"Не удалось открыть папку: {e}")
    
    def export_to_csv(self, filepath: str) -> Tuple[bool, str]:
        """
        Экспорт в CSV файл
        Returns: (success, error_message)
        """
        try:
            # Проверка возможности записи в директорию
            directory = os.path.dirname(filepath)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Проверка прав на запись
            if os.path.exists(filepath):
                if not os.access(filepath, os.W_OK):
                    return False, f"Нет прав на запись в файл: {filepath}"
            else:
                if directory and not os.access(directory, os.W_OK):
                    return False, f"Нет прав на запись в директорию: {directory}"
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                writer.writerow(['#' * 80])
                writer.writerow(['# ADAPTIVE EDGE STORAGE SIMULATOR - РЕЗУЛЬТАТЫ СИМУЛЯЦИИ'])
                writer.writerow(['#' * 80])
                writer.writerow([])
                
                writer.writerow(['[ПАРАМЕТРЫ СИМУЛЯЦИИ]'])
                writer.writerow(['Параметр', 'Значение'])
                for key, value in self.simulation_params.items():
                    writer.writerow([key, value])
                writer.writerow(['Время начала', self.simulation_start_time])
                writer.writerow(['Время окончания', self.simulation_end_time])
                writer.writerow([])
                
                summary = self.get_summary()
                writer.writerow(['[СВОДНАЯ СТАТИСТИКА]'])
                writer.writerow(['Метрика', 'Значение'])
                for key, value in summary.items():
                    writer.writerow([key, value])
                writer.writerow([])
                
                writer.writerow(['[ДОСТУПНОСТЬ ДАННЫХ ПО ВРЕМЕНИ]'])
                writer.writerow(['Время', 'Доступность (0-1)'])
                for time, score in self.availability_history:
                    writer.writerow([f'{time:.2f}', f'{score:.4f}'])
                writer.writerow([])
                
                writer.writerow(['[СОСТОЯНИЕ УЗЛОВ]'])
                writer.writerow(['Время', 'Работающие узлы', 'Всего узлов'])
                for time, online, total in self.health_history:
                    writer.writerow([f'{time:.2f}', online, total])
                writer.writerow([])
                
                writer.writerow(['[ОТКАЗЫ УЗЛОВ]'])
                writer.writerow(['Время', 'ID узла'])
                for time, node_id in self.node_failures:
                    writer.writerow([f'{time:.2f}', node_id])
                writer.writerow([])
                
                writer.writerow(['[ОПЕРАЦИИ ЗАПИСИ]'])
                writer.writerow(['ID блока', 'Записано реплик', 'Целевое число реплик'])
                for block_id, written, target in self.write_successes:
                    writer.writerow([block_id, written, target])
                writer.writerow([])
                
                writer.writerow(['[ОПЕРАЦИИ ВОССТАНОВЛЕНИЯ]'])
                writer.writerow(['ID блока', 'Узел-источник', 'Узел-назначение'])
                for block_id, source, target in self.repair_successes:
                    writer.writerow([block_id, source, target])
            
            self._open_file_location(filepath)
            return True, ""
            
        except PermissionError as e:
            return False, f"Ошибка доступа: {e}. Возможно, файл открыт в другой программе."
        except OSError as e:
            return False, f"Ошибка файловой системы: {e}"
        except Exception as e:
            return False, f"Неизвестная ошибка: {e}"
    
    def export_to_excel(self, filepath: str) -> Tuple[bool, str]:
        """
        Экспорт в Excel файл с форматированием и графиками
        Returns: (success, error_message)
        """
        if not OPENPYXL_AVAILABLE:
            return False, "Библиотека openpyxl не установлена. Установите: pip install openpyxl"
        
        try:
            # Проверка возможности записи в директорию
            directory = os.path.dirname(filepath)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Проверка прав на запись
            if os.path.exists(filepath):
                if not os.access(filepath, os.W_OK):
                    return False, f"Нет прав на запись в файл: {filepath}"
            else:
                if directory and not os.access(directory, os.W_OK):
                    return False, f"Нет прав на запись в директорию: {directory}"
            
            wb = Workbook()
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(bold=True, size=14)
            
            # ===== Лист 1: Параметры и сводка =====
            ws_summary = wb.active
            ws_summary.title = "Параметры и сводка"
            
            # Заголовок
            ws_summary['A1'] = "АДАПТИВНЫЙ EDGE STORAGE SIMULATOR"
            ws_summary['A1'].font = title_font
            ws_summary.merge_cells('A1:C1')
            
            # Параметры симуляции
            ws_summary['A3'] = "ПАРАМЕТРЫ СИМУЛЯЦИИ"
            ws_summary['A3'].font = header_font
            ws_summary['A3'].fill = header_fill
            
            row = 4
            for key, value in self.simulation_params.items():
                ws_summary[f'A{row}'] = str(key)
                ws_summary[f'B{row}'] = str(value)
                row += 1
            ws_summary[f'A{row}'] = "Время начала"
            ws_summary[f'B{row}'] = str(self.simulation_start_time)
            row += 1
            ws_summary[f'A{row}'] = "Время окончания"
            ws_summary[f'B{row}'] = str(self.simulation_end_time)
            row += 2
            
            # Сводная статистика
            ws_summary[f'A{row}'] = "СВОДНАЯ СТАТИСТИКА"
            ws_summary[f'A{row}'].font = header_font
            ws_summary[f'A{row}'].fill = header_fill
            row += 1
            
            summary = self.get_summary()
            for key, value in summary.items():
                ws_summary[f'A{row}'] = str(key)
                ws_summary[f'B{row}'] = value
                row += 1
            
            # ===== Лист 2: Доступность данных =====
            ws_avail = wb.create_sheet("Доступность данных")
            ws_avail['A1'] = "Время"
            ws_avail['B1'] = "Доступность"
            ws_avail['A1'].font = header_font
            ws_avail['B1'].font = header_font
            ws_avail['A1'].fill = header_fill
            ws_avail['B1'].fill = header_fill
            
            for i, (time, score) in enumerate(self.availability_history, start=2):
                ws_avail[f'A{i}'] = time
                ws_avail[f'B{i}'] = score
            
            # Добавление графика доступности
            if len(self.availability_history) > 1:
                chart = LineChart()
                chart.title = "Доступность данных во времени"
                chart.x_axis.title = "Время"
                chart.y_axis.title = "Доступность"
                
                data = Reference(ws_avail, min_col=2, min_row=1, max_row=len(self.availability_history)+1)
                categories = Reference(ws_avail, min_col=1, min_row=2, max_row=len(self.availability_history)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                ws_avail.add_chart(chart, "D2")
            
            # ===== Лист 3: Состояние узлов =====
            ws_health = wb.create_sheet("Состояние узлов")
            ws_health['A1'] = "Время"
            ws_health['B1'] = "Работающие узлы"
            ws_health['C1'] = "Всего узлов"
            for col in ['A', 'B', 'C']:
                ws_health[f'{col}1'].font = header_font
                ws_health[f'{col}1'].fill = header_fill
            
            for i, (time, online, total) in enumerate(self.health_history, start=2):
                ws_health[f'A{i}'] = time
                ws_health[f'B{i}'] = online
                ws_health[f'C{i}'] = total
            
            # График состояния узлов
            if len(self.health_history) > 1:
                chart = LineChart()
                chart.title = "Состояние узлов во времени"
                chart.x_axis.title = "Время"
                chart.y_axis.title = "Количество узлов"
                
                data = Reference(ws_health, min_col=2, max_col=3, min_row=1, max_row=len(self.health_history)+1)
                categories = Reference(ws_health, min_col=1, min_row=2, max_row=len(self.health_history)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                ws_health.add_chart(chart, "E2")
            
            # ===== Лист 4: Отказы узлов =====
            ws_failures = wb.create_sheet("Отказы узлов")
            ws_failures['A1'] = "Время"
            ws_failures['B1'] = "ID узла"
            ws_failures['A1'].font = header_font
            ws_failures['B1'].font = header_font
            ws_failures['A1'].fill = header_fill
            ws_failures['B1'].fill = header_fill
            
            for i, (time, node_id) in enumerate(self.node_failures, start=2):
                ws_failures[f'A{i}'] = time
                ws_failures[f'B{i}'] = node_id
            
            # ===== Лист 5: Операции =====
            ws_ops = wb.create_sheet("Операции")
            ws_ops['A1'] = "Тип операции"
            ws_ops['B1'] = "ID блока"
            ws_ops['C1'] = "Доп. информация"
            for col in ['A', 'B', 'C']:
                ws_ops[f'{col}1'].font = header_font
                ws_ops[f'{col}1'].fill = header_fill
            
            row = 2
            for block_id, written, target in self.write_successes:
                ws_ops[f'A{row}'] = "Запись успешна"
                ws_ops[f'B{row}'] = block_id
                ws_ops[f'C{row}'] = f"{written}/{target} реплик"
                row += 1
            
            for block_id, reason in self.write_failures:
                ws_ops[f'A{row}'] = "Запись не удалась"
                ws_ops[f'B{row}'] = block_id
                ws_ops[f'C{row}'] = reason
                row += 1
            
            for block_id, source, target in self.repair_successes:
                ws_ops[f'A{row}'] = "Восстановление"
                ws_ops[f'B{row}'] = block_id
                ws_ops[f'C{row}'] = f"с {source} на {target}"
                row += 1
            
            # ===== Автоматическая ширина столбцов (ИСПРАВЛЕНО) =====
            for ws in wb.worksheets:
                # Получаем все столбцы, которые используются в листе
                if ws.max_row > 0 and ws.max_column > 0:
                    for col_idx in range(1, ws.max_column + 1):
                        max_length = 0
                        col_letter = None
                        try:
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                        except AttributeError:
                            # Если не удалось получить column_letter, пропускаем
                            continue
                        
                        # Проходим по строкам в этом столбце
                        for row_idx in range(1, ws.max_row + 1):
                            try:
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if cell.value is not None:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except Exception:
                                continue
                        
                        # Устанавливаем ширину с запасом
                        adjusted_width = min(max_length + 2, 50)
                        if col_letter:
                            ws.column_dimensions[col_letter].width = adjusted_width
            
            wb.save(filepath)
            self._open_file_location(filepath)
            return True, ""
            
        except PermissionError as e:
            return False, f"Ошибка доступа: {e}. Возможно, файл открыт в Excel. Закройте файл и повторите попытку."
        except OSError as e:
            return False, f"Ошибка файловой системы: {e}"
        except Exception as e:
            return False, f"Неизвестная ошибка: {e}"